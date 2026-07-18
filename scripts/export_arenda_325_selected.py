import argparse
import math
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent.parent  # project root
SCRIPTS_DIR = Path(__file__).resolve().parent
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

import etagi_commerce_scraper as scraper  # noqa: E402


BASE_URL = "https://spb.etagi.com"

TYPE_LABELS = {
    18: "Офис",
    19: "Торговое помещение",
    20: "Помещение свободного назначения",
    23: "Свободное назначение",
}

TYPE_SLUG_LABELS = {
    "office": "Офис",
    "commercial": "Торговое помещение",
    "other": "Свободное назначение",
    "base": "База",
}

REPAIR_LABELS = {
    2: "Косметический ремонт",
    6: "Современный ремонт",
    7: "Евро",
    9: "Требует ремонта",
    10: "Черновая отделка",
    11: "Улучшенная черновая отделка",
    13: "Предчистовая отделка",
    15: "С отделкой",
    19: "Чистовая отделка",
    24: "Дизайнерский / евро / современный ремонт",
}

LAYOUT_LABELS = {
    1: "Кабинетная",
    2: "Свободная",
    3: "Смешанная",
}

WATER_SOURCE_LABELS = {
    1: "Централизованный",
    2: "Автономный",
}

LOCATION_LABELS = {
    1: "БЦ / ТЦ / ТРК",
    2: "В нежилом здании",
    3: "Отдельностоящее здание целиком",
    4: "В жилом здании",
    5: "Пристрой в жилом доме",
}

PARKING_LABELS = {
    1: "Общественная",
    2: "Собственная",
    3: "В здании",
}


HEADERS = [
    "Код объекта",
    "Заголовок",
    "Тип",
    "url",
    "Адрес",
    "До центра",
    "Метро",
    "ЖК",
    "Статус ЖК",
    "Цена",
    "Старая цена",
    "Цена за м²",
    "Этаж / Этажность",
    "Общая площадь",
    "Отдельный вход",
    "Количество входов",
    "Мощность электричества",
    "Наличие вытяжки",
    "Высота потолков",
    "Планировка",
    "Ремонт",
    "Источник воды",
    "Вентиляция",
    "Кондиционер",
    "Пожарная сигнализация",
    "Расположение",
    "Год постройки",
    "Стены",
    "Парковка",
]


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return ILLEGAL_CHARACTERS_RE.sub("", scraper.clean_text(value))


def number(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace(".", ",")


def int_value(value: Any) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def nested_get(obj: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        current: Any = obj
        found = True
        for part in key.split("."):
            if not isinstance(current, dict) or part not in current:
                found = False
                break
            current = current[part]
        if found and current not in (None, "", []):
            return current
    return None


def yes_no(value: Any) -> str:
    if isinstance(value, bool):
        return "Да" if value else "Нет"
    if isinstance(value, (int, float)) and value in {0, 1}:
        return "Да" if value == 1 else "Нет"
    if isinstance(value, str):
        lowered = value.strip().lower()
        if lowered in {"true", "1", "yes", "да"}:
            return "Да"
        if lowered in {"false", "0", "no", "нет"}:
            return "Нет"
    return clean(value)


def map_value(value: Any, mapping: dict[int, str]) -> str:
    key = int_value(value)
    if key is not None:
        return mapping.get(key, clean(value))
    return clean(value)


def rub(value: Any, suffix: str = "") -> str:
    if value in (None, ""):
        return ""
    try:
        text = f"{int(float(value)):,}".replace(",", " ")
    except (TypeError, ValueError):
        text = clean(value)
    return f"{text}{suffix}"


def square(value: Any) -> str:
    value_text = number(value)
    return f"{value_text} м²" if value_text else ""


def kwt(value: Any) -> str:
    value_text = number(value)
    return f"{value_text} кВт" if value_text else ""


def meters(value: Any) -> str:
    value_text = number(value)
    return f"{value_text} м" if value_text else ""


def km(value: Any) -> str:
    value_text = number(value)
    return f"{value_text} км" if value_text else ""


def type_label(obj: dict[str, Any]) -> str:
    object_type_id = int_value(obj.get("object_type_id"))
    if object_type_id in TYPE_LABELS:
        return TYPE_LABELS[object_type_id]
    return TYPE_SLUG_LABELS.get(clean(obj.get("type")), clean(obj.get("type")))


def title(obj: dict[str, Any]) -> str:
    label = type_label(obj)
    area = number(obj.get("square"))
    return f"{label}, {area}м²" if label and area else label


def street_with_prefix(street: str) -> str:
    lowered = street.lower()
    no_prefix_tokens = [
        "проспект",
        "шоссе",
        "площадь",
        "линия",
        "переулок",
        "набережная",
        "бульвар",
        "аллея",
    ]
    if any(token in lowered for token in no_prefix_tokens):
        return street
    return f"ул. {street}" if street else ""


def address(obj: dict[str, Any]) -> str:
    meta = obj.get("meta") or {}
    parts = []
    city = clean(meta.get("city"))
    street = street_with_prefix(clean(meta.get("street")))
    house = clean(obj.get("house_address_number") or obj.get("house_num"))
    if city:
        parts.append(city)
    if street:
        parts.append(street)
    if house:
        parts.append(house)
    text = ", ".join(parts)
    center = clean(obj.get("to_center"))
    if center:
        text = f"{text} ({center} км до центра)"
    return text


def metro_lines(obj: dict[str, Any]) -> str:
    stations = obj.get("metro_stations") or []
    rows: list[str] = []
    for index, station in enumerate(stations):
        name = clean(station.get("stationName") or station.get("name")).removeprefix("метро ").strip()
        distance = station.get("distance")
        time = clean(obj.get("time_to_metro")) if index == 0 else ""
        if not time and distance:
            time = str(max(1, math.floor(float(distance) / 66.0)))
        if name and time and distance:
            rows.append(f"{name} - {time} мин. ({float(distance) / 1000:g} км)")
        elif name and distance:
            rows.append(f"{name} ({float(distance) / 1000:g} км)")
        elif name:
            rows.append(name)
    if rows:
        return "\n".join(rows)
    name = clean(obj.get("metro_station"))
    time = clean(obj.get("time_to_metro"))
    distance = obj.get("metro_distance")
    if name and time and distance:
        return f"{name} - {time} мин. ({float(distance) / 1000:g} км)"
    return name


def residential_complex_status(obj: dict[str, Any]) -> str:
    year = int_value(obj.get("building_year") or obj.get("deadline_y"))
    if year and year <= 2026:
        return "Сдан"
    if year:
        quarter = clean(obj.get("deadline_q"))
        return f"{quarter} кв. {year}" if quarter else str(year)
    return ""


def detail_object(path: Path) -> tuple[dict[str, Any], dict[str, str]]:
    html = path.read_text(encoding="utf-8")
    data = scraper.load_page_data(html)
    return data.get("objects", {}).get("commerceObject") or {}, scraper.extract_characteristics(html)


def cached_detail_files(detail_dir: Path) -> dict[str, Path]:
    files: dict[str, Path] = {}
    for path in sorted(detail_dir.glob("*.html")):
        object_id = path.stem.split("_", 1)[0]
        if object_id and object_id not in files:
            files[object_id] = path
    return files


def list_object_ids(list_dir: Path) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for path in sorted(list_dir.glob("*.html")):
        data = scraper.load_page_data(path.read_text(encoding="utf-8"))
        for item in data.get("lists", {}).get("commerce") or []:
            object_id = clean(scraper.object_id(item))
            if object_id and object_id not in seen:
                seen.add(object_id)
                result.append(object_id)
    return result


def characteristic(chars: dict[str, str], key: str) -> str:
    return clean(chars.get(key))


def build_row(obj: dict[str, Any], chars: dict[str, str]) -> dict[str, str]:
    meta = obj.get("meta") or {}
    additional = obj.get("additional_meta") or {}

    floor = characteristic(chars, "Этаж / Этажность")
    if not floor:
        floor_num = clean(additional.get("floor") or obj.get("floor"))
        floors = clean(additional.get("floors") or obj.get("floors"))
        floor = f"{floor_num} из {floors}" if floor_num and floors else floor_num

    area = characteristic(chars, "Общая площадь") or square(obj.get("square"))
    entrance_count = characteristic(chars, "Количество входов") or clean(additional.get("entrance_cnt"))
    exhaust = characteristic(chars, "Наличие вытяжки") or yes_no(
        nested_get(obj, "additional_meta.vent", "vent")
    )
    layout = characteristic(chars, "Планировка") or map_value(additional.get("layout"), LAYOUT_LABELS)

    year = clean(obj.get("building_year") or obj.get("deadline_y"))

    return {
        "Код объекта": clean(obj.get("object_id") or obj.get("_ticket_id")),
        "Заголовок": title(obj),
        "Тип": type_label(obj),
        "url": f"{BASE_URL}/commerce/{clean(obj.get('object_id') or obj.get('_ticket_id'))}/",
        "Адрес": address(obj),
        "До центра": km(obj.get("to_center")),
        "Метро": metro_lines(obj),
        "ЖК": clean(meta.get("newcomplex")),
        "Статус ЖК": residential_complex_status(obj),
        "Цена": rub(obj.get("price"), " /мес."),
        "Старая цена": rub(obj.get("old_price"), " /мес."),
        "Цена за м²": rub(obj.get("price_m2"), " /мес. за м²"),
        "Этаж / Этажность": floor,
        "Общая площадь": area,
        "Отдельный вход": yes_no(
            nested_get(obj, "separate_entrance", "additional_meta.private_entrance")
        ),
        "Количество входов": entrance_count,
        "Мощность электричества": kwt(additional.get("power")),
        "Наличие вытяжки": exhaust,
        "Высота потолков": characteristic(chars, "Высота потолков") or meters(additional.get("ceilings_height")),
        "Планировка": layout,
        "Ремонт": map_value(obj.get("keep"), REPAIR_LABELS),
        "Источник воды": map_value(additional.get("water_sources") or obj.get("waterSource"), WATER_SOURCE_LABELS),
        "Вентиляция": yes_no(nested_get(obj, "additional_meta.ventilation", "ventilation")),
        "Кондиционер": yes_no(nested_get(obj, "additional_meta.conditioner", "conditioner")),
        "Пожарная сигнализация": yes_no(nested_get(obj, "additional_meta.fireAlarm", "fireAlarm")),
        "Расположение": map_value(additional.get("location") or obj.get("location"), LOCATION_LABELS),
        "Год постройки": year,
        "Стены": clean(meta.get("walls")),
        "Парковка": map_value(additional.get("parkings"), PARKING_LABELS),
    }


def build_rows(cache_dir: Path) -> list[dict[str, str]]:
    list_dir = cache_dir / "list"
    detail_dir = cache_dir / "detail"
    detail_files = cached_detail_files(detail_dir)
    rows = []
    for object_id in list_object_ids(list_dir):
        detail_path = detail_files.get(object_id)
        if not detail_path:
            continue
        obj, chars = detail_object(detail_path)
        rows.append(build_row(obj, chars))
    return rows


def save_excel(rows: list[dict[str, str]], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "selected"
    ws.append(HEADERS)
    for row in rows:
        ws.append([row.get(header, "") for header in HEADERS])

    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = {
        "A": 14,
        "B": 32,
        "C": 28,
        "D": 34,
        "E": 52,
        "F": 14,
        "G": 45,
        "H": 28,
        "I": 14,
        "J": 16,
        "K": 16,
        "L": 20,
    }
    for col_idx in range(1, len(HEADERS) + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = widths.get(letter, 22)

    wb.save(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export selected visible Etagi arenda fields.")
    parser.add_argument("--cache-dir", default=str(ROOT / "raw" / "arenda_325" / "html_cache"))
    parser.add_argument("--output", default=str(ROOT / "raw" / "arenda_325" / "etagi_arenda_325_selected.xlsx"))
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    rows = build_rows(Path(args.cache_dir))
    if not rows:
        raise RuntimeError("No rows exported.")
    save_excel(rows, Path(args.output))
    print(f"Saved {len(rows)} rows to {args.output}")


if __name__ == "__main__":
    main()
