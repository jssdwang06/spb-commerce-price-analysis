import argparse
import math
import random
import re
import time
from datetime import datetime
from hashlib import sha1
from pathlib import Path
from typing import Any
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


BASE_URL = "https://spb.etagi.com"
FILTER_PARAMS = [
    ("m_obj_type[]", "19"),
    ("m_obj_type[]", "23"),
    ("m_obj_type[]", "18"),
    ("m_obj_type[]", "20"),
]
DEFAULT_LIST_PATH = "/commerce/"
LIST_URL = f"{BASE_URL}{DEFAULT_LIST_PATH}"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/126.0.0.0 Safari/537.36"
    ),
    "Accept": (
        "text/html,application/xhtml+xml,application/xml;q=0.9,"
        "image/avif,image/webp,image/apng,*/*;q=0.8,"
        "application/signed-exchange;v=b3;q=0.7"
    ),
    "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
    "Accept-Encoding": "gzip, deflate, br",
    "Cache-Control": "max-age=0",
    "Connection": "keep-alive",
    "DNT": "1",
    "Referer": f"{BASE_URL}/commerce/",
    "Sec-CH-UA": '"Not/A)Brand";v="8", "Chromium";v="126", "Google Chrome";v="126"',
    "Sec-CH-UA-Mobile": "?0",
    "Sec-CH-UA-Platform": '"Windows"',
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "same-origin",
    "Sec-Fetch-User": "?1",
    "Upgrade-Insecure-Requests": "1",
}

BASE_COLUMNS = [
    "object_id",
    "url",
    "Тип",
    "Площадь",
    "Цена",
    "Цена за м²",
    "Метро",
    "Адрес",
    "Ремонт",
    "Отопление",
    "Вентиляция",
    "Кондиционер",
    "Канализация / cанузлы",
    "Видеонаблюдение",
    "Пожарная сигнализация",
    "Линия",
    "Расположение",
    "Год постройки",
    "Стены",
    "Парковка",
    "Мебель",
    "detail_error",
]

DETAIL_SECTION_TITLES = {"Характеристики", "О здании", "Коммуникации"}
SKIP_CHARACTERISTICS = {
    "Код объекта",
    "Отдельный вход",
    "Мощность электричества",
    "Высота потолков",
}

REPAIR_MAP = {
    1: "Строительная отделка",
    2: "Косметический ремонт",
    5: "Частично требуется ремонт",
    6: "Современный ремонт",
    7: "Евро",
    8: "Дизайн",
    9: "Требуется ремонт",
    10: "Черновая отделка",
    11: "Улучшенная черновая отделка",
    14: "Предчистовая отделка",
    15: "С отделкой",
    24: "Дизайн / Евро / Современный ремонт",
}

HEAT_SUPPLY_MAP = {
    1: "Центральное",
    2: "Газовое",
    3: "Электрическое",
    4: "Твердотопливное",
    5: "Отсутствует",
}

SEWERAGE_MAP = {
    1: "В помещении",
    2: "Общественная в здании",
    3: "На этаже",
}

LINE_MAP = {
    1: "Первая",
    2: "Вторая",
    3: "Третья",
}

LOCATION_MAP = {
    1: "Отдельное здание",
    2: "В бизнес-центре",
    3: "Отдельно стоящее",
    4: "В жилом здании",
}

WALL_MAP = {
    83: "Кирпичные",
}

PARKING_MAP = {
    1: "Общественная",
    2: "Собственная",
    3: "В здании",
}


def load_headers() -> dict[str, str]:
    headers = dict(HEADERS)
    try:
        from headers import headers as custom_headers
    except ImportError:
        return headers
    if not isinstance(custom_headers, dict):
        raise TypeError("headers.py must define a dict named headers")
    headers.update({str(key): str(value) for key, value in custom_headers.items()})
    return headers


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def rub(value: Any) -> str:
    if value in (None, ""):
        return ""
    try:
        return f"{int(float(value)):,}".replace(",", " ") + " ₽"
    except (TypeError, ValueError):
        return clean_text(value)


def square(value: Any) -> str:
    if value in (None, ""):
        return ""
    number = str(value).replace(".", ",")
    return f"{number} м²"


def load_page_data(html: str) -> dict[str, Any]:
    soup = BeautifulSoup(html, "html.parser")
    for script in soup.find_all("script"):
        text = script.string or script.get_text() or ""
        if text.startswith("var data="):
            return requests.models.complexjson.loads(text[len("var data="):].rstrip(";"))
    raise ValueError("Cannot find embedded var data JSON")


def cache_path(cache_dir: str, kind: str, key: str, url: str) -> Path | None:
    if not cache_dir:
        return None
    safe_key = re.sub(r"[^0-9A-Za-z_.-]+", "_", str(key)).strip("_")
    digest = sha1(url.encode("utf-8")).hexdigest()[:10]
    return Path(cache_dir) / kind / f"{safe_key}_{digest}.html"


def has_html_cache(cache_dir: str, kind: str, key: str, url: str) -> bool:
    path = cache_path(cache_dir, kind, key, url)
    return bool(path and path.exists())


def request_html(
    session: requests.Session,
    url: str,
    retries: int = 3,
    debug_http: bool = False,
    cache_file: Path | None = None,
    use_cache: bool = False,
    offline_cache: bool = False,
) -> str:
    if cache_file and (use_cache or offline_cache) and cache_file.exists():
        print(f"HTML cache hit: {cache_file}")
        return cache_file.read_text(encoding="utf-8")
    if offline_cache:
        raise FileNotFoundError(f"HTML cache miss: {cache_file}")

    last_error: Exception | None = None
    for attempt in range(1, retries + 1):
        try:
            response = session.get(url, timeout=35)
            if response.status_code in {403, 429}:
                if debug_http:
                    debug_path = f"debug_http_{response.status_code}_{int(time.time())}.html"
                    with open(debug_path, "w", encoding="utf-8") as file:
                        file.write(response.text)
                    print(f"HTTP debug status={response.status_code} final_url={response.url}")
                    print(f"HTTP debug headers={dict(response.headers)}")
                    print(f"HTTP debug body saved to {debug_path}")
                raise requests.HTTPError(f"{response.status_code} blocked", response=response)
            response.raise_for_status()
            if "captcha" in response.url.lower():
                raise RuntimeError(f"Captcha page returned: {response.url}")
            if cache_file:
                cache_file.parent.mkdir(parents=True, exist_ok=True)
                cache_file.write_text(response.text, encoding="utf-8")
                print(f"HTML cached: {cache_file}")
            return response.text
        except requests.RequestException as exc:
            last_error = exc
            wait = random.uniform(10, 20) * attempt
            print(f"Request failed ({attempt}/{retries}): {url}; sleep {wait:.1f}s")
            time.sleep(wait)
    raise RuntimeError(f"Failed to fetch {url}") from last_error


def normalize_list_path(value: str) -> str:
    path = clean_text(value) or DEFAULT_LIST_PATH
    if path.startswith(BASE_URL):
        path = path[len(BASE_URL) :]
    if not path.startswith("/"):
        path = "/" + path
    if not path.endswith("/"):
        path += "/"
    return path


def set_list_path(value: str) -> None:
    global LIST_URL
    LIST_URL = f"{BASE_URL}{normalize_list_path(value)}"


def list_url(page: int) -> str:
    params = [f"{key}={value}" for key, value in FILTER_PARAMS]
    if page > 1:
        params.append(f"page={page}")
    return f"{LIST_URL}?{'&'.join(params)}"


def object_id(obj: dict[str, Any]) -> Any:
    return obj.get("object_id") or obj.get("_ticket_id")


def type_from_page_data(data: dict[str, Any], obj: dict[str, Any]) -> str:
    by_slug = {
        "office": "Офис",
        "commercial": "Торговое помещение",
        "base": "База",
        "other": "Свободное назначение",
    }
    raw_type = clean_text(obj.get("type"))
    if raw_type in by_slug:
        return by_slug[raw_type]

    breadcrumbs = (
        data.get("injectedSettings", {})
        .get("serverBreadCrumbs", {})
    )
    for rows in breadcrumbs.values():
        for row in rows:
            url = row.get("url", "")
            title = clean_text(row.get("title"))
            if title and "type[]=" in url and title != "Коммерческая недвижимость":
                return title

    object_types = data.get("objects", {}).get("objectTypes") or {}
    object_type = clean_text(object_types.get("objectType_a") or object_types.get("objectType_g"))
    object_type = re.sub(r"^недвижимость\s+", "", object_type, flags=re.IGNORECASE)
    if object_type:
        return object_type[:1].upper() + object_type[1:]

    by_id = {
        18: "Офис",
        19: "Торговое помещение",
        20: "Помещение свободного назначения",
        23: "Свободное назначение",
    }
    return by_id.get(int(obj.get("object_type_id") or 0), raw_type)


def format_metro(obj: dict[str, Any]) -> str:
    stations = obj.get("metro_stations") or []
    names = []
    for station in stations:
        name = clean_text(station.get("stationName") or station.get("name"))
        name = re.sub(r"^метро\s+", "", name, flags=re.IGNORECASE)
        if name and name not in names:
            names.append(name)
    if names:
        return ", ".join(names)
    return clean_text(obj.get("metro_station"))


def format_address(obj: dict[str, Any]) -> str:
    meta = obj.get("meta") or {}
    parts = []
    city = clean_text(meta.get("city"))
    street = clean_text(meta.get("street"))
    house = clean_text(obj.get("house_address_number") or obj.get("house_num"))
    if city:
        parts.append(city)
    if street:
        parts.append(street)
    if house:
        parts.append(house)
    return ", ".join(parts)


def list_pairs(container: Any) -> dict[str, str]:
    pairs: dict[str, str] = {}
    for li in container.find_all("li"):
        parts = [clean_text(part) for part in li.stripped_strings]
        parts = [part for part in parts if part]
        if len(parts) < 2:
            continue
        key = parts[0]
        value = " ".join(parts[1:])
        if key in SKIP_CHARACTERISTICS:
            continue
        if key and value and key not in DETAIL_SECTION_TITLES and key not in pairs:
            pairs[key] = value
    return pairs


def find_section_container(title_tag: Any) -> Any:
    current = title_tag
    for _ in range(6):
        if current is None:
            break
        if len(list_pairs(current)) >= 1:
            return current
        current = current.parent
    return title_tag.parent


def extract_characteristics(html: str) -> dict[str, str]:
    soup = BeautifulSoup(html, "html.parser")
    result: dict[str, str] = {}
    for title in DETAIL_SECTION_TITLES:
        for title_tag in soup.find_all(string=lambda text: clean_text(text) == title):
            container = find_section_container(title_tag.parent)
            for key, value in list_pairs(container).items():
                if key not in result:
                    result[key] = value
    return result


def is_present(value: Any) -> bool:
    return value is not None and value != "" and value != [] and value != {}


def first_present(obj: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        value = obj.get(key)
        if is_present(value):
            return value
    return None


def format_bool(value: Any) -> str:
    if isinstance(value, bool):
        return "Да" if value else "Нет"
    if isinstance(value, str):
        lowered = value.strip().lower()
        if lowered in {"t", "true", "1", "yes", "да"}:
            return "Да"
        if lowered in {"f", "false", "0", "no", "нет"}:
            return "Нет"
    if isinstance(value, (int, float)) and value in {0, 1}:
        return "Да" if value == 1 else "Нет"
    return clean_text(value)


def format_presence(value: Any) -> str:
    if isinstance(value, bool):
        return "Есть" if value else "Нет"
    return format_bool(value)


def format_mapped(value: Any, mapping: dict[int, str]) -> str:
    if not is_present(value):
        return ""
    if isinstance(value, list):
        parts = [format_mapped(item, mapping) for item in value]
        return " / ".join(part for part in parts if part)
    try:
        key = int(value)
    except (TypeError, ValueError):
        return clean_text(value)
    return mapping.get(key, clean_text(value))


def detail_fields_from_object(obj: dict[str, Any]) -> dict[str, str]:
    additional_meta = obj.get("additional_meta") or {}
    meta = obj.get("meta") or {}
    combined = {**additional_meta, **obj}
    fields = {
        "Ремонт": format_mapped(first_present(combined, "keep", "repair"), REPAIR_MAP),
        "Отопление": format_mapped(first_present(combined, "heatSupply", "heat_supply"), HEAT_SUPPLY_MAP),
        "Вентиляция": format_bool(first_present(combined, "ventilation", "vent")),
        "Кондиционер": format_bool(first_present(combined, "conditioner")),
        "Канализация / cанузлы": format_mapped(first_present(combined, "sewerage"), SEWERAGE_MAP),
        "Видеонаблюдение": format_bool(first_present(combined, "videoSurveillance", "video", "remoteSecurity")),
        "Пожарная сигнализация": format_bool(first_present(combined, "fireAlarm")),
        "Линия": format_mapped(first_present(combined, "line"), LINE_MAP),
        "Расположение": format_mapped(first_present(combined, "location"), LOCATION_MAP),
        "Год постройки": clean_text(first_present(combined, "building_year", "yearConstruct")),
        "Стены": clean_text(meta.get("walls")) or format_mapped(first_present(combined, "wall_id", "wall"), WALL_MAP),
        "Парковка": format_mapped(first_present(combined, "parkings", "parking"), PARKING_MAP),
        "Мебель": format_presence(first_present(combined, "furniture")),
    }
    return {key: value for key, value in fields.items() if value}


def base_row(data: dict[str, Any], obj: dict[str, Any]) -> dict[str, Any]:
    oid = object_id(obj)
    return {
        "object_id": oid,
        "url": f"{BASE_URL}/commerce/{oid}/" if oid else "",
        "Тип": type_from_page_data(data, obj),
        "Площадь": square(obj.get("square")),
        "Цена": rub(obj.get("price")),
        "Цена за м²": rub(obj.get("price_m2")) + "/м²" if obj.get("price_m2") else "",
        "Метро": format_metro(obj),
        "Адрес": format_address(obj),
        "detail_error": "",
    }


def collect_list_page(
    session: requests.Session,
    page: int,
    debug_http: bool = False,
    cache_dir: str = "",
    use_cache: bool = False,
    offline_cache: bool = False,
) -> tuple[list[dict[str, Any]], dict[str, Any], int]:
    url = list_url(page)
    html = request_html(
        session,
        url,
        debug_http=debug_http,
        cache_file=cache_path(cache_dir, "list", f"page_{page:04d}", url),
        use_cache=use_cache,
        offline_cache=offline_cache,
    )
    data = load_page_data(html)
    items = data.get("lists", {}).get("commerce") or []
    total = int(data.get("filters", {}).get("commerce", {}).get("count") or len(items))
    print(f"List page {page}: {len(items)} items, total={total}, url={url}")
    return items, data, total


def collect_detail(
    session: requests.Session,
    oid: Any,
    debug_http: bool = False,
    cache_dir: str = "",
    use_cache: bool = False,
    offline_cache: bool = False,
) -> tuple[dict[str, Any], dict[str, Any]]:
    url = f"{BASE_URL}/commerce/{oid}/"
    html = request_html(
        session,
        url,
        debug_http=debug_http,
        cache_file=cache_path(cache_dir, "detail", str(oid), url),
        use_cache=use_cache,
        offline_cache=offline_cache,
    )
    data = load_page_data(html)
    chars = extract_characteristics(html)
    detail_obj = data.get("objects", {}).get("commerceObject") or {}
    for key, value in detail_fields_from_object(detail_obj).items():
        if not chars.get(key):
            chars[key] = value
    print(f"Detail {oid}: {len(chars)} characteristics")
    return data, chars


def ordered_columns(rows: list[dict[str, Any]]) -> list[str]:
    columns = list(BASE_COLUMNS)
    for row in rows:
        for key in row:
            if key not in columns:
                columns.append(key)
    return columns


def save_excel(rows: list[dict[str, Any]], path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "commerce"
    columns = ordered_columns(rows)
    ws.append(columns)
    for row in rows:
        ws.append([row.get(column, "") for column in columns])

    ws.freeze_panes = "A2"
    for col_idx, column in enumerate(columns, start=1):
        width = max(12, min(45, len(column) + 2))
        for cell in ws[get_column_letter(col_idx)][1: min(len(rows) + 1, 30)]:
            width = max(width, min(45, len(str(cell.value or "")) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(path)


def load_existing_rows(path: str) -> list[dict[str, Any]]:
    try:
        wb = load_workbook(path)
    except FileNotFoundError:
        return []
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {headers[index]: value for index, value in enumerate(values)}
        if row.get("object_id"):
            rows.append(row)
    return rows


def sleep_between(delay: tuple[float, float], label: str) -> None:
    wait = random.uniform(*delay)
    print(f"{label}: sleep {wait:.1f}s")
    time.sleep(wait)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Scrape filtered spb.etagi.com commerce listings to Excel.")
    parser.add_argument(
        "--list-path",
        default=DEFAULT_LIST_PATH,
        help="Listing path or full listing URL, for example /commerce/arenda/.",
    )
    parser.add_argument("--pages", type=int, default=1, help="Number of list pages to scrape.")
    parser.add_argument("--all", action="store_true", help="Scrape all pages from the filtered total count.")
    parser.add_argument("--output", default="", help="Output .xlsx path.")
    parser.add_argument("--resume", action="store_true", help="Continue an existing output file.")
    parser.add_argument("--save-every", type=int, default=1, help="Save after every N rows.")
    parser.add_argument("--continue-on-error", action="store_true", help="Keep list row when detail page fails.")
    parser.add_argument("--detail-delay-min", type=float, default=8.0, help="Minimum delay between detail pages.")
    parser.add_argument("--detail-delay-max", type=float, default=18.0, help="Maximum delay between detail pages.")
    parser.add_argument("--page-delay-min", type=float, default=20.0, help="Minimum delay between list pages.")
    parser.add_argument("--page-delay-max", type=float, default=45.0, help="Maximum delay between list pages.")
    parser.add_argument("--debug-http", action="store_true", help="Save blocked HTTP response bodies for inspection.")
    parser.add_argument("--html-cache-dir", default="", help="Directory for saved list/detail HTML pages.")
    parser.add_argument("--use-html-cache", action="store_true", help="Read saved HTML before requesting the site.")
    parser.add_argument("--offline-cache", action="store_true", help="Use only saved HTML; fail if a page is missing.")
    parser.add_argument("--download-html-only", action="store_true", help="Only save list/detail HTML pages; do not write Excel.")
    parser.add_argument("--download-list-only", action="store_true", help="Only save list HTML pages; do not download details or write Excel.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    set_list_path(args.list_path)
    if args.detail_delay_min > args.detail_delay_max:
        raise ValueError("--detail-delay-min cannot exceed --detail-delay-max")
    if args.page_delay_min > args.page_delay_max:
        raise ValueError("--page-delay-min cannot exceed --page-delay-max")
    if (args.use_html_cache or args.offline_cache) and not args.html_cache_dir:
        raise ValueError("--use-html-cache and --offline-cache require --html-cache-dir")
    if args.download_html_only and not args.html_cache_dir:
        raise ValueError("--download-html-only requires --html-cache-dir")
    use_html_cache = args.use_html_cache or args.offline_cache

    session = requests.Session()
    session.headers.update(load_headers())

    if not args.output:
        suffix = "filtered_all" if args.all else f"filtered_page{args.pages}"
        args.output = f"etagi_commerce_{suffix}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"

    rows = load_existing_rows(args.output) if args.resume else []
    existing_ids = {str(row.get("object_id")) for row in rows if row.get("object_id")}
    if existing_ids:
        print(f"Resume: loaded {len(existing_ids)} existing rows from {args.output}")

    all_items: list[tuple[dict[str, Any], dict[str, Any]]] = []
    first_items, first_data, total_count = collect_list_page(
        session,
        1,
        debug_http=args.debug_http,
        cache_dir=args.html_cache_dir,
        use_cache=use_html_cache,
        offline_cache=args.offline_cache,
    )
    all_items.extend((item, first_data) for item in first_items)

    total_pages = math.ceil(total_count / 30) if args.all else args.pages
    total_pages = max(1, total_pages)
    print(f"Planned pages: {total_pages}; filtered total reported by site: {total_count}")

    for page in range(2, total_pages + 1):
        list_cached = has_html_cache(args.html_cache_dir, "list", f"page_{page:04d}", list_url(page))
        if not args.offline_cache and not (use_html_cache and list_cached):
            sleep_between((args.page_delay_min, args.page_delay_max), f"Before list page {page}")
        items, data, total_count = collect_list_page(
            session,
            page,
            debug_http=args.debug_http,
            cache_dir=args.html_cache_dir,
            use_cache=use_html_cache,
            offline_cache=args.offline_cache,
        )
        all_items.extend((item, data) for item in items)

    if args.download_list_only:
        print(f"List HTML download finished: {args.html_cache_dir}")
        return

    seen = set()
    unique_items: list[tuple[dict[str, Any], dict[str, Any]]] = []
    for item, data in all_items:
        oid = object_id(item)
        if oid and oid not in seen:
            seen.add(oid)
            unique_items.append((item, data))

    if args.download_html_only:
        print(f"Download HTML only: {len(unique_items)} detail pages")
        for index, (item, _list_data) in enumerate(unique_items, start=1):
            oid = object_id(item)
            detail_url = f"{BASE_URL}/commerce/{oid}/"
            if use_html_cache and has_html_cache(args.html_cache_dir, "detail", str(oid), detail_url):
                print(f"[{index}/{len(unique_items)}] detail HTML {oid}: already cached, skip")
                continue
            print(f"[{index}/{len(unique_items)}] download detail HTML {oid}")
            try:
                collect_detail(
                    session,
                    oid,
                    debug_http=args.debug_http,
                    cache_dir=args.html_cache_dir,
                    use_cache=use_html_cache,
                    offline_cache=args.offline_cache,
                )
            except Exception as exc:
                if not args.continue_on_error:
                    raise
                print(f"Detail {oid}: failed: {clean_text(exc)}")
            if index < len(unique_items) and not args.offline_cache:
                sleep_between((args.detail_delay_min, args.detail_delay_max), f"After detail {oid}")
        print(f"HTML download finished: {args.html_cache_dir}")
        return

    for index, (item, list_data) in enumerate(unique_items, start=1):
        row = base_row(list_data, item)
        oid = row["object_id"]
        if str(oid) in existing_ids:
            print(f"[{index}/{len(unique_items)}] {oid}: already saved, skip")
            continue

        print(f"[{index}/{len(unique_items)}] {oid}")
        try:
            detail_data, characteristics = collect_detail(
                session,
                oid,
                debug_http=args.debug_http,
                cache_dir=args.html_cache_dir,
                use_cache=use_html_cache,
                offline_cache=args.offline_cache,
            )
            detail_obj = detail_data.get("objects", {}).get("commerceObject") or {}
            row.update(base_row(detail_data, detail_obj or item))
            row.update(characteristics)
        except Exception as exc:
            if not args.continue_on_error:
                save_excel(rows, args.output)
                raise
            row["detail_error"] = clean_text(exc)
            print(f"Detail {oid}: failed: {row['detail_error']}")

        rows.append(row)
        existing_ids.add(str(oid))
        if args.save_every > 0 and len(rows) % args.save_every == 0:
            save_excel(rows, args.output)
            print(f"Checkpoint saved {len(rows)} rows to {args.output}")

        if index < len(unique_items) and not args.offline_cache:
            sleep_between((args.detail_delay_min, args.detail_delay_max), f"After detail {oid}")

    save_excel(rows, args.output)
    print(f"Saved {len(rows)} rows to {args.output}")


if __name__ == "__main__":
    main()
